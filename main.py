import re
from openpyxl import Workbook


def redeTxt(url):
    # 返回List，每个元素包含一个题的字符串
    with open(url, "r", encoding="gbk") as f:
        str = f.read().strip().split("[P]")
    for i in range(len(str)):
        str[i] = str[i].strip()
    return str[:-1]


def handleTxt(questionStr):
    # 返回题目列表
    patternI = re.compile(r'\[I\].*')
    patternQ = re.compile(r'\[Q\].*')
    patternA = re.compile(r'\[A\].*')
    patternB = re.compile(r'\[B\].*')
    patternC = re.compile(r'\[C\].*')
    patternD = re.compile(r'\[D\].*')

    questionList = []
    for i in questionStr:
        one = []
        one.append(re.search(patternI, i).group()[3:])
        one.append(re.search(patternQ, i).group()[3:])
        one.append(re.search(patternA, i).group()[3:])
        one.append(re.search(patternB, i).group()[3:])
        one.append(re.search(patternC, i).group()[3:])
        one.append(re.search(patternD, i).group()[3:])
        questionList.append(one)

    return questionList

def writeExcel(questionList):
    # 初始化工作薄
    book = Workbook()

    # 设置工作表
    # sheet = book.create_sheet('工作表 1', 0)
    sheet = book.active

    for i in range(len(questionList)):
        sheet.cell(row=i + 1, column=1).value = questionList[i][1]
        sheet.cell(row=i + 1, column=2).value = 'A'
        sheet.cell(row=i + 1, column=3).value = questionList[i][0]
        sheet.cell(row=i + 1, column=4).value = questionList[i][2]
        sheet.cell(row=i + 1, column=5).value = questionList[i][3]
        sheet.cell(row=i + 1, column=6).value = questionList[i][4]
        sheet.cell(row=i + 1, column=7).value = questionList[i][5]


    book.save('导题.xlsx')


url = r".\A类题库(v20211022).txt"
writeExcel(handleTxt(redeTxt(url)))

